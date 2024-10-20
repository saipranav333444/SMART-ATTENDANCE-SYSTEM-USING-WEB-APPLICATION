from flask import Flask, render_template, request, redirect, url_for, Response, send_from_directory,jsonify
import cv2
import face_recognition
from openpyxl import Workbook
from datetime import date
import os
import time
import keyboard
import sys
from moviepy.editor import VideoFileClip
import csv
from werkzeug.utils import secure_filename
from urllib.parse import unquote
import numpy as np
import mysql.connector
# Specify the path where the video will be stored
image_path = r'C:\Users\HP\Desktop\FRS\classroom'

app = Flask(__name__, static_url_path='/templates')
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="root",
  database="asish"
)

def login_check(username, password):
    mycursor=mydb.cursor()
    query="SELECT * FROM users WHERE username=%s AND password=%s"
    mycursor.execute(query,(username,password))
    myresult=mycursor.fetchall()
    if myresult:
        if myresult[0][2]=="Y":
            return True,"admin"
        else:
            return True,"not"
    else:
        return False,"Invalid username or password"
def generate_frames(duration, timee):
    known_face_encodings = []
    known_face_names = []
    temp=[]
    def load_known_faces(folder_path,csv_path):

        print("Press 'q' to exit or wait for 5 minutes.")
        for file_name in os.listdir(folder_path):
            if file_name.endswith(('.jpg', '.png', '.jpeg')):
                name = os.path.splitext(file_name)[0]
                image_path = os.path.join(folder_path, file_name)
                face_image = face_recognition.load_image_file(image_path)
                face_encoding = face_recognition.face_encodings(face_image)[0]

                known_face_encodings.append(face_encoding)
                known_face_names.append(name)
        flattened_arrays = [arr.flatten() for arr in known_face_encodings]
        try:
            with open(csv_path, 'w', newline='') as csvfile:
                csv_writer = csv.writer(csvfile)
                csv_writer.writerow(known_face_names)
                csv_writer.writerows(flattened_arrays)
        except Exception as e:
            print(f"An error occurred: {e}")
    def read_known_faces(csv_path):
        with open(csv_path, 'r') as csvfile:
            csv_reader = csv.reader(csvfile)
            known_face_names.extend(next(csv_reader))
            print(known_face_names)
            for row in csv_reader:
                face_encoding = np.array(row).astype(float)
                known_face_encodings.append(face_encoding)
    timee=timee.upper()
    known_faces_folder = 'C:/Users/HP/Desktop/FRS/'+timee
    csv_path = 'C:/Users/HP/Desktop/FRS/csv/'+timee+'.csv'
    if(os.path.exists(csv_path)):
        read_known_faces(csv_path)
    else:
        load_known_faces(known_faces_folder,csv_path)
    # Load known faces
    print(known_face_encodings,known_face_names)
    # Initialize some variables
    face_names = []
    already_attendance_taken = set()

    # Open or create workbook and sheet
    wb_path = 'C:/Users/HP/Desktop/FRS/attendance_excel_'
    wb = Workbook()
    sheet1 = wb.active
    #print("a")
    inp = timee+'_'+duration
    sheet1.title = inp if inp else 'Sheet1'  # Set default title if no input provided
    wb_path = wb_path + inp + '.xlsx'
    sheet1['A1'] = 'Name/Date'
    sheet1['B1'] = str(date.today())
    row = 2
    col = 1
    x=len(known_face_names)
    temp_names=known_face_names.copy()
    temp_names.sort()
    temp_att=[0]*x
    print(temp_names,temp_att)
    #print("b")
    try:
        start_time = time.time()
        
        #print("c")
        for image_filename in os.listdir(image_path):  # Assuming image_path is the path where you store uploaded images
        # Read the image
            #print("d")
            frame = cv2.imread(os.path.join(image_path, image_filename))
            #print("e")
            print(image_filename)
            #print("f")
        # Find all face locations and face encodings in the current image
            face_locations = face_recognition.face_locations(frame, model="hog")
            face_encodings = face_recognition.face_encodings(frame, face_locations)

            face_names = []
            for face_encoding, face_location in zip(face_encodings, face_locations):
            # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.6)
                name = "Unknown" if not any(matches) else known_face_names[matches.index(True)]

                face_names.append(name)
                if (name not in already_attendance_taken) and (name != "Unknown"):
                    ind=temp_names.index(name)
                    temp_att[ind]=1
                    already_attendance_taken.add(name)
        # Display the results
            for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Draw a rectangle around the face
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
                cv2.rectangle(frame, (left, bottom + 20), (right, bottom), (0, 0, 255), cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom + 16), font, 0.5, (255, 255, 255), 1)

        # Display the resulting image
            cv2.imshow('Image', frame)
            cv2.waitKey(1)
            os.remove(os.path.join(image_path, image_filename))
        for i in range(x):
            sheet1.cell(row=row, column=col, value=temp_names[i])
            col += 1
            if temp_att[i]==1:
                sheet1.cell(row=row, column=col, value="Present")
            else:
                sheet1.cell(row=row, column=col, value="Absent")
            row += 1
            col = 1
        pst=temp_att.count(1)
        ast=temp_att.count(0)
        countt=len(temp_att)
        row+=2
        sheet1.cell(row=row,column=col,value="No of Students Present")
        col+=1
        sheet1.cell(row=row,column=col,value=str(pst))
        col+=2
        sheet1.cell(row=row,column=col,value="Presentees Percentage")
        col+=1
        pp=float(pst)/float(countt)
        pp*=100
        pp=round(pp,2)
        tpp=str(pp)+'%'
        sheet1.cell(row=row,column=col,value=tpp)
        
        col=1
        row+=1
        sheet1.cell(row=row,column=col,value="No of Students Absent")
        col+=1
        sheet1.cell(row=row,column=col,value=str(ast))
        col+=2
        sheet1.cell(row=row,column=col,value="Absentees Percentage")
        col+=1
        ap=float(ast)/float(countt)
        ap*=100
        ap=round(ap,2)
        abpp=str(ap)+'%'
        sheet1.cell(row=row,column=col,value=abpp)
        
        print("Attendance Taken")
            

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Save workbook to the specified path
        wb_path = os.path.join('templates', 'attendance_excel_' + inp + '.xlsx')
        wb.save(wb_path)
        cv2.destroyAllWindows()

def create_folder(folder_path):
    try:
        os.makedirs(folder_path, exist_ok=True)
        return True, f"Folder created successfully."
    except Exception as e:
        return False, f"Error creating folder: {str(e)}"

def save_uploaded_photos(folder_path, files):
    try:
        for file in files:
            if file.filename != '':
                filename = os.path.join(folder_path, file.filename)
                if not os.path.exists(filename):
                    file.save(filename)
        
        return True, "Photos uploaded successfully."
    except Exception as e:
        return False, f"Error uploading photos: {str(e)}"
def delete_photos(folder_path, files_to_delete):
    try:
        for file_to_delete in files_to_delete:
            file_path = os.path.join(folder_path, file_to_delete)
            if os.path.exists(file_path):
                os.remove(file_path)
        return True, "Photos deleted successfully."
    except Exception as e:
        return False, f"An error occurred: {e}"

@app.route('/delete_photos', methods=['POST'])
def delete_photos_route():
    folder_name = request.form['delete_folder_name']
    folder_name=folder_name.upper()
    folder_path = f'C:/Users/HP/Desktop/FRS/{folder_name}'
    files_to_delete = request.form.getlist('delete_files[]')

    success, message = delete_photos(folder_path, files_to_delete)

    return render_template('admin.html', folder_message="", del_photos_message=message, folder_name=folder_name)
@app.route('/login', methods=['POST'])
def login_page():
    username = request.form['username']
    password = request.form['password']

    login_status, message = login_check(username, password)

    if login_status:
        # If login successful, redirect to the video feed page
        if message=="admin":
            return redirect(url_for('admin'))
        else:
           return redirect(url_for('index'))
    else:
        # If login unsuccessful, render login page with error message
        return render_template('login.html', login_message=message, login_status="error")

# Login page route
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')
# Index page route
@app.route('/')
def login():
    return render_template('login.html')



@app.route('/start_attendance', methods=['POST'])
def start_attendance():
    duration = request.form['duration']
    timee = request.form['timee']

    generate_frames(duration, timee)
    # For example, you can pass them to a function for processing
    # start_attendance_process(duration, timee)

    return "Attendance process started for Duration: {} and Time: {}".format(duration, timee)

@app.route('/download/<string:filename>')
def download(filename):
    return send_from_directory('templates', filename)

@app.route('/create_folder', methods=['POST'])
def create_folder_route():
    folder_name = request.form.get('folderName')
    folder_name=folder_name.upper()
    if not folder_name:
        return render_template('admin.html', folder_message="Folder name is required.", photos_message="")

    folder_path = f'C:/Users/HP/Desktop/FRS/{folder_name}'
    if os.path.exists(folder_path):
        message="Folder Already exists"
    else:
        success, message = create_folder(folder_path)

    return render_template('admin.html', folder_message=message, photos_message="")

@app.route('/upload_photos', methods=['POST'])
def upload_photos():
    folder_name = request.form['folder_name']
    folder_name=folder_name.upper()
    tempp= f'C:/Users/HP/Desktop/FRS/csv/{folder_name}.csv'
    folder_path = f'C:/Users/HP/Desktop/FRS/{folder_name}'
    files = request.files.getlist('files[]')

    if not files or all(file.filename == '' for file in files):
        return render_template('admin.html', folder_message="", photos_message="Please select photos to upload.")

    success, message = save_uploaded_photos(folder_path, files)
    if os.path.exists(tempp):
        os.remove(tempp)

    return render_template('admin.html', folder_message="", photos_message=message, folder_name=folder_name)

@app.route('/upload', methods=['POST'])
def upload():
    # Check if the post request has the file part
    if 'files[]' not in request.files:
        return redirect(url_for('index'))

    files = request.files.getlist('files[]')

    # If the user submits an empty form
    if not files or all(file.filename == '' for file in files):
        return redirect(url_for('index'))

    # Save each file to the specified path
    for file in files:
        if file.filename != '':
            filename = secure_filename(file.filename)
            file.save(os.path.join(image_path, filename))

    return redirect(url_for('index'))
    


if __name__ == "__main__":
    app.run(debug=False, host='0.0.0.0', port=5000)